"""Excel/Spreadsheet MCP Server Implementation.

Provides Excel and spreadsheet operations through MCP protocol:
- Read/write Excel workbooks
- Sheet manipulation
- Cell operations
- Formula evaluation
- Pivot table creation
- Table extraction
- Financial report formatting
- Financial model parsing (LBO, DCF)
"""

import logging
from pathlib import Path
from typing import Any, Optional

try:
    import openpyxl
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    openpyxl = None

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    pd = None

logger = logging.getLogger(__name__)


class ExcelMCPServer:
    """Excel/Spreadsheet MCP server implementation."""

    def __init__(self, config: dict[str, Any]):
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl library not installed. "
                "Install with: pip install openpyxl"
            )
        
        self.config = config
        self.max_rows = config.get("max_rows", 100000)
        self.max_columns = config.get("max_columns", 1000)
        self.evaluate_formulas = config.get("evaluate_formulas", True)

    async def read_workbook(
        self,
        excel_path: str,
        data_only: bool = False,
    ) -> dict[str, Any]:
        """Read Excel workbook metadata.

        Args:
            excel_path: Path to Excel file
            data_only: Read data only (no formulas)

        Returns:
            Workbook information
        """
        try:
            path = Path(excel_path)
            if not path.exists():
                return {
                    "success": False,
                    "error": f"Excel file not found: {excel_path}",
                }
            
            wb = openpyxl.load_workbook(path, data_only=data_only)
            
            sheets_info = []
            for sheet in wb.worksheets:
                sheets_info.append({
                    "name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "dimensions": sheet.dimensions,
                })
            
            return {
                "success": True,
                "excel_path": str(path),
                "sheets": sheets_info,
                "sheet_count": len(wb.worksheets),
                "active_sheet": wb.active.title,
            }

        except Exception as e:
            logger.error(f"Failed to read workbook {excel_path}: {e}")
            return {
                "success": False,
                "error": f"Failed to read workbook: {str(e)}",
                "excel_path": excel_path,
            }

    async def write_workbook(
        self,
        excel_path: str,
        sheets: dict[str, list[list[Any]]],
        overwrite: bool = False,
    ) -> dict[str, Any]:
        """Write data to Excel workbook.

        Args:
            excel_path: Path to Excel file
            sheets: Dictionary of sheet_name -> data (list of rows)
            overwrite: Overwrite existing file

        Returns:
            Write result
        """
        try:
            path = Path(excel_path)
            if path.exists() and not overwrite:
                return {
                    "success": False,
                    "error": f"File already exists: {excel_path}. Set overwrite=True to overwrite.",
                }
            
            wb = openpyxl.Workbook()
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            for sheet_name, data in sheets.items():
                ws = wb.create_sheet(sheet_name)
                
                for row_idx, row_data in enumerate(data, start=1):
                    for col_idx, cell_value in enumerate(row_data, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=cell_value)
            
            wb.save(path)
            
            return {
                "success": True,
                "excel_path": str(path),
                "sheets_written": list(sheets.keys()),
                "sheet_count": len(sheets),
            }

        except Exception as e:
            logger.error(f"Failed to write workbook {excel_path}: {e}")
            return {
                "success": False,
                "error": f"Failed to write workbook: {str(e)}",
                "excel_path": excel_path,
            }

    async def read_sheet(
        self,
        excel_path: str,
        sheet_name: Optional[str] = None,
        range_spec: Optional[str] = None,
        as_dataframe: bool = False,
    ) -> dict[str, Any]:
        """Read specific sheet from Excel.

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name (None for active sheet)
            range_spec: Cell range (e.g., "A1:D10", None for all)
            as_dataframe: Return as pandas DataFrame

        Returns:
            Sheet data
        """
        try:
            path = Path(excel_path)
            if not path.exists():
                return {
                    "success": False,
                    "error": f"Excel file not found: {excel_path}",
                }
            
            wb = openpyxl.load_workbook(path, data_only=True)
            
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    return {
                        "success": False,
                        "error": f"Sheet '{sheet_name}' not found",
                        "available_sheets": wb.sheetnames,
                    }
                ws = wb[sheet_name]
            else:
                ws = wb.active
            
            # Get data range
            if range_spec:
                cell_range = ws[range_spec]
                if isinstance(cell_range, tuple):
                    # Multiple rows
                    data = [[cell.value for cell in row] for row in cell_range]
                else:
                    # Single cell or row
                    data = [[cell.value for cell in cell_range]]
            else:
                # All data
                data = [[cell.value for cell in row] for row in ws.iter_rows()]
            
            # Convert to DataFrame if requested
            if as_dataframe and PANDAS_AVAILABLE:
                df = pd.DataFrame(data[1:], columns=data[0] if data else None)
                return {
                    "success": True,
                    "excel_path": str(path),
                    "sheet_name": ws.title,
                    "data": df.to_dict(orient="records"),
                    "columns": df.columns.tolist(),
                    "row_count": len(df),
                }
            
            return {
                "success": True,
                "excel_path": str(path),
                "sheet_name": ws.title,
                "data": data,
                "row_count": len(data),
                "column_count": len(data[0]) if data else 0,
            }

        except Exception as e:
            logger.error(f"Failed to read sheet from {excel_path}: {e}")
            return {
                "success": False,
                "error": f"Failed to read sheet: {str(e)}",
                "excel_path": excel_path,
            }

    async def get_cell_value(
        self,
        excel_path: str,
        sheet_name: str,
        cell_address: str,
    ) -> dict[str, Any]:
        """Get value of specific cell.

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name
            cell_address: Cell address (e.g., "A1", "B5")

        Returns:
            Cell value
        """
        try:
            path = Path(excel_path)
            if not path.exists():
                return {
                    "success": False,
                    "error": f"Excel file not found: {excel_path}",
                }
            
            wb = openpyxl.load_workbook(path, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' not found",
                }
            
            ws = wb[sheet_name]
            cell = ws[cell_address]
            
            return {
                "success": True,
                "excel_path": str(path),
                "sheet_name": sheet_name,
                "cell_address": cell_address,
                "value": cell.value,
                "data_type": str(type(cell.value).__name__),
            }

        except Exception as e:
            logger.error(f"Failed to get cell value from {excel_path}: {e}")
            return {
                "success": False,
                "error": f"Failed to get cell value: {str(e)}",
                "excel_path": excel_path,
            }

    async def set_cell_value(
        self,
        excel_path: str,
        sheet_name: str,
        cell_address: str,
        value: Any,
    ) -> dict[str, Any]:
        """Set value of specific cell.

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name
            cell_address: Cell address (e.g., "A1")
            value: Value to set

        Returns:
            Operation result
        """
        try:
            path = Path(excel_path)
            if not path.exists():
                return {
                    "success": False,
                    "error": f"Excel file not found: {excel_path}",
                }
            
            wb = openpyxl.load_workbook(path)
            
            if sheet_name not in wb.sheetnames:
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' not found",
                }
            
            ws = wb[sheet_name]
            ws[cell_address] = value
            
            wb.save(path)
            
            return {
                "success": True,
                "excel_path": str(path),
                "sheet_name": sheet_name,
                "cell_address": cell_address,
                "value": value,
            }

        except Exception as e:
            logger.error(f"Failed to set cell value in {excel_path}: {e}")
            return {
                "success": False,
                "error": f"Failed to set cell value: {str(e)}",
                "excel_path": excel_path,
            }

    async def evaluate_formula(
        self,
        formula: str,
        context: Optional[dict[str, Any]] = None,
    ) -> dict[str, Any]:
        """Evaluate Excel formula.

        Args:
            formula: Excel formula (e.g., "=SUM(A1:A10)")
            context: Variable context for formula

        Returns:
            Formula result
        """
        try:
            # Simple formula evaluation (limited support)
            # This is a placeholder - full Excel formula evaluation requires additional libraries
            
            formula = formula.strip()
            if formula.startswith("="):
                formula = formula[1:]
            
            # Basic operations
            if formula.startswith("SUM("):
                # Extract range
                import re
                match = re.search(r'SUM\((.*?)\)', formula)
                if match and context:
                    range_ref = match.group(1)
                    values = context.get(range_ref, [])
                    result = sum(values)
                    return {
                        "success": True,
                        "formula": formula,
                        "result": result,
                    }
            
            return {
                "success": False,
                "error": "Formula evaluation not fully implemented. Use openpyxl data_only mode for calculated values.",
                "formula": formula,
            }

        except Exception as e:
            logger.error(f"Failed to evaluate formula: {e}")
            return {
                "success": False,
                "error": f"Failed to evaluate formula: {str(e)}",
                "formula": formula,
            }

    async def create_pivot(
        self,
        excel_path: str,
        source_sheet: str,
        dest_sheet: str,
        source_range: str,
        pivot_config: dict[str, Any],
    ) -> dict[str, Any]:
        """Create pivot table (using pandas).

        Args:
            excel_path: Path to Excel file
            source_sheet: Source sheet name
            dest_sheet: Destination sheet name
            source_range: Source data range
            pivot_config: Pivot configuration (index, columns, values, aggfunc)

        Returns:
            Pivot result
        """
        if not PANDAS_AVAILABLE:
            return {
                "success": False,
                "error": "pandas not available. Install with: pip install pandas",
            }
        
        try:
            # Read source data
            result = await self.read_sheet(excel_path, source_sheet, source_range, as_dataframe=True)
            if not result["success"]:
                return result
            
            df = pd.DataFrame(result["data"])
            
            # Create pivot table
            pivot = pd.pivot_table(
                df,
                index=pivot_config.get("index"),
                columns=pivot_config.get("columns"),
                values=pivot_config.get("values"),
                aggfunc=pivot_config.get("aggfunc", "sum"),
            )
            
            # Write back to Excel
            path = Path(excel_path)
            with pd.ExcelWriter(path, engine='openpyxl', mode='a') as writer:
                pivot.to_excel(writer, sheet_name=dest_sheet)
            
            return {
                "success": True,
                "excel_path": str(path),
                "source_sheet": source_sheet,
                "dest_sheet": dest_sheet,
                "pivot_shape": pivot.shape,
            }

        except Exception as e:
            logger.error(f"Failed to create pivot table: {e}")
            return {
                "success": False,
                "error": f"Failed to create pivot table: {str(e)}",
            }

    async def extract_tables(
        self,
        excel_path: str,
        sheet_name: Optional[str] = None,
    ) -> dict[str, Any]:
        """Extract data tables from Excel sheet.

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name (None for active)

        Returns:
            Extracted tables
        """
        try:
            # Read sheet data
            result = await self.read_sheet(excel_path, sheet_name)
            if not result["success"]:
                return result
            
            data = result["data"]
            
            # Simple table detection: find blocks of non-empty cells
            tables = []
            current_table = []
            empty_row_count = 0
            
            for row in data:
                if any(cell is not None for cell in row):
                    current_table.append(row)
                    empty_row_count = 0
                else:
                    empty_row_count += 1
                    if empty_row_count > 2 and current_table:
                        # End of table
                        tables.append(current_table)
                        current_table = []
            
            if current_table:
                tables.append(current_table)
            
            return {
                "success": True,
                "excel_path": excel_path,
                "sheet_name": result["sheet_name"],
                "tables": tables,
                "table_count": len(tables),
            }

        except Exception as e:
            logger.error(f"Failed to extract tables from {excel_path}: {e}")
            return {
                "success": False,
                "error": f"Failed to extract tables: {str(e)}",
                "excel_path": excel_path,
            }

    async def format_financial_report(
        self,
        excel_path: str,
        sheet_name: str,
        title: str,
        data: dict[str, Any],
    ) -> dict[str, Any]:
        """Generate formatted financial report.

        Args:
            excel_path: Path to Excel file
            sheet_name: Sheet name
            title: Report title
            data: Financial data dictionary

        Returns:
            Formatting result
        """
        try:
            path = Path(excel_path)
            
            # Create or load workbook
            if path.exists():
                wb = openpyxl.load_workbook(path)
            else:
                wb = openpyxl.Workbook()
            
            # Create or get sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(sheet_name)
            
            # Add title
            ws['A1'] = title
            ws['A1'].font = Font(bold=True, size=14)
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Add data
            row = 3
            for key, value in data.items():
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=1).font = Font(bold=True)
                
                # Format numbers
                if isinstance(value, (int, float)):
                    ws.cell(row=row, column=2).number_format = '#,##0.00'
                
                row += 1
            
            # Auto-size columns
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 15
            
            wb.save(path)
            
            return {
                "success": True,
                "excel_path": str(path),
                "sheet_name": sheet_name,
                "title": title,
                "rows_written": len(data),
            }

        except Exception as e:
            logger.error(f"Failed to format financial report: {e}")
            return {
                "success": False,
                "error": f"Failed to format report: {str(e)}",
            }

    async def parse_financial_model(
        self,
        excel_path: str,
        model_type: str = "dcf",
    ) -> dict[str, Any]:
        """Parse financial model (LBO, DCF, etc.).

        Args:
            excel_path: Path to Excel file
            model_type: Model type (dcf, lbo, merger)

        Returns:
            Parsed model data
        """
        try:
            path = Path(excel_path)
            if not path.exists():
                return {
                    "success": False,
                    "error": f"Excel file not found: {excel_path}",
                }
            
            wb = openpyxl.load_workbook(path, data_only=True)
            
            # Parse based on model type
            parsed_data = {
                "model_type": model_type,
                "sheets": [],
            }
            
            for sheet in wb.worksheets:
                sheet_data = {
                    "name": sheet.title,
                    "key_metrics": {},
                }
                
                # Look for common financial terms
                financial_keywords = [
                    "revenue", "ebitda", "capex", "fcf", "wacc",
                    "terminal value", "npv", "irr", "equity value"
                ]
                
                for row in sheet.iter_rows(values_only=True):
                    if row[0]:
                        row_label = str(row[0]).lower()
                        for keyword in financial_keywords:
                            if keyword in row_label:
                                # Extract numeric value from row
                                for cell in row[1:]:
                                    if isinstance(cell, (int, float)) and cell != 0:
                                        sheet_data["key_metrics"][keyword] = cell
                                        break
                
                if sheet_data["key_metrics"]:
                    parsed_data["sheets"].append(sheet_data)
            
            return {
                "success": True,
                "excel_path": str(path),
                "model_type": model_type,
                "parsed_data": parsed_data,
                "sheets_analyzed": len(parsed_data["sheets"]),
            }

        except Exception as e:
            logger.error(f"Failed to parse financial model {excel_path}: {e}")
            return {
                "success": False,
                "error": f"Failed to parse model: {str(e)}",
                "excel_path": excel_path,
            }


def get_server_definition() -> dict[str, Any]:
    """Get Excel MCP server definition.

    Returns:
        Server definition dictionary
    """
    return {
        "name": "excel",
        "category": "documents",
        "description": "Excel/spreadsheet operations (read, write, formulas, financial models)",
        "tools": [
            {
                "name": "read_workbook",
                "description": "Read Excel workbook metadata and sheets",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "excel_path": {
                            "type": "string",
                            "description": "Path to Excel file",
                        },
                        "data_only": {
                            "type": "boolean",
                            "description": "Read data only (no formulas)",
                            "default": False,
                        },
                    },
                    "required": ["excel_path"],
                },
            },
            {
                "name": "write_workbook",
                "description": "Write data to Excel workbook",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "excel_path": {
                            "type": "string",
                            "description": "Path to Excel file",
                        },
                        "sheets": {
                            "type": "object",
                            "description": "Dictionary of sheet_name -> data (list of rows)",
                        },
                        "overwrite": {
                            "type": "boolean",
                            "description": "Overwrite existing file",
                            "default": False,
                        },
                    },
                    "required": ["excel_path", "sheets"],
                },
            },
            {
                "name": "read_sheet",
                "description": "Read specific sheet from Excel",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "excel_path": {
                            "type": "string",
                            "description": "Path to Excel file",
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Sheet name (omit for active sheet)",
                        },
                        "range_spec": {
                            "type": "string",
                            "description": "Cell range (e.g., 'A1:D10', omit for all)",
                        },
                        "as_dataframe": {
                            "type": "boolean",
                            "description": "Return as pandas DataFrame",
                            "default": False,
                        },
                    },
                    "required": ["excel_path"],
                },
            },
            {
                "name": "get_cell_value",
                "description": "Get value of specific cell",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "excel_path": {
                            "type": "string",
                            "description": "Path to Excel file",
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Sheet name",
                        },
                        "cell_address": {
                            "type": "string",
                            "description": "Cell address (e.g., 'A1', 'B5')",
                        },
                    },
                    "required": ["excel_path", "sheet_name", "cell_address"],
                },
            },
            {
                "name": "set_cell_value",
                "description": "Set value of specific cell",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "excel_path": {
                            "type": "string",
                            "description": "Path to Excel file",
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Sheet name",
                        },
                        "cell_address": {
                            "type": "string",
                            "description": "Cell address (e.g., 'A1')",
                        },
                        "value": {
                            "description": "Value to set (string, number, formula)",
                        },
                    },
                    "required": ["excel_path", "sheet_name", "cell_address", "value"],
                },
            },
            {
                "name": "evaluate_formula",
                "description": "Evaluate Excel formula",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "formula": {
                            "type": "string",
                            "description": "Excel formula (e.g., '=SUM(A1:A10)')",
                        },
                        "context": {
                            "type": "object",
                            "description": "Variable context for formula",
                        },
                    },
                    "required": ["formula"],
                },
            },
            {
                "name": "create_pivot",
                "description": "Create pivot table from data",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "excel_path": {
                            "type": "string",
                            "description": "Path to Excel file",
                        },
                        "source_sheet": {
                            "type": "string",
                            "description": "Source sheet name",
                        },
                        "dest_sheet": {
                            "type": "string",
                            "description": "Destination sheet name",
                        },
                        "source_range": {
                            "type": "string",
                            "description": "Source data range",
                        },
                        "pivot_config": {
                            "type": "object",
                            "description": "Pivot configuration (index, columns, values, aggfunc)",
                        },
                    },
                    "required": ["excel_path", "source_sheet", "dest_sheet", "source_range", "pivot_config"],
                },
            },
            {
                "name": "extract_tables",
                "description": "Extract data tables from Excel sheet",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "excel_path": {
                            "type": "string",
                            "description": "Path to Excel file",
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Sheet name (omit for active)",
                        },
                    },
                    "required": ["excel_path"],
                },
            },
            {
                "name": "format_financial_report",
                "description": "Generate formatted financial report",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "excel_path": {
                            "type": "string",
                            "description": "Path to Excel file",
                        },
                        "sheet_name": {
                            "type": "string",
                            "description": "Sheet name",
                        },
                        "title": {
                            "type": "string",
                            "description": "Report title",
                        },
                        "data": {
                            "type": "object",
                            "description": "Financial data dictionary",
                        },
                    },
                    "required": ["excel_path", "sheet_name", "title", "data"],
                },
            },
            {
                "name": "parse_financial_model",
                "description": "Parse financial model (LBO, DCF, etc.)",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "excel_path": {
                            "type": "string",
                            "description": "Path to Excel file",
                        },
                        "model_type": {
                            "type": "string",
                            "enum": ["dcf", "lbo", "merger"],
                            "description": "Model type",
                            "default": "dcf",
                        },
                    },
                    "required": ["excel_path"],
                },
            },
        ],
        "resources": [],
        "metadata": {
            "version": "1.0.0",
            "priority": "high",
            "category": "documents",
            "requires": [
                "openpyxl>=3.1.2",
                "pandas>=2.0.0",
                "xlrd>=2.0.1",
            ],
            "performance_target": "<500ms for read, <1s for write",
        },
    }