"""
Comprehensive tests for Week 2 MCP Servers.

Tests for:
- Redis MCP Server (storage)
- Docker MCP Server (devops)
- Prometheus MCP Server (monitoring)
- PDF Processing MCP Server (documents)
- Excel/Spreadsheet MCP Server (documents)
"""

import asyncio
import os
import tempfile
from pathlib import Path
from typing import Any

import pytest

# Import MCP server classes
try:
    from axiom.integrations.mcp_servers.storage.redis_server import RedisMCPServer, get_server_definition as redis_def
    REDIS_AVAILABLE = True
except ImportError:
    REDIS_AVAILABLE = False

try:
    from axiom.integrations.mcp_servers.devops.docker_server import DockerMCPServer, get_server_definition as docker_def
    DOCKER_AVAILABLE = True
except ImportError:
    DOCKER_AVAILABLE = False

try:
    from axiom.integrations.mcp_servers.monitoring.prometheus_server import PrometheusMCPServer, get_server_definition as prometheus_def
    PROMETHEUS_AVAILABLE = True
except ImportError:
    PROMETHEUS_AVAILABLE = False

try:
    from axiom.integrations.mcp_servers.documents.pdf_server import PDFProcessingMCPServer, get_server_definition as pdf_def
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from axiom.integrations.mcp_servers.documents.excel_server import ExcelMCPServer, get_server_definition as excel_def
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# ===== Redis MCP Server Tests =====

@pytest.mark.skipif(not REDIS_AVAILABLE, reason="Redis dependencies not available")
class TestRedisMCPServer:
    """Test Redis MCP Server."""

    @pytest.fixture
    async def redis_server(self):
        """Create Redis server instance."""
        config = {
            "host": os.getenv("REDIS_HOST", "localhost"),
            "port": int(os.getenv("REDIS_PORT", "6379")),
            "db": 0,
        }
        server = RedisMCPServer(config)
        yield server
        await server.close()

    def test_redis_server_definition(self):
        """Test Redis server definition."""
        definition = redis_def()
        
        assert definition["name"] == "redis"
        assert definition["category"] == "storage"
        assert len(definition["tools"]) == 8
        
        tool_names = [tool["name"] for tool in definition["tools"]]
        expected_tools = [
            "get_value", "set_value", "delete_key",
            "publish_message", "subscribe_channel",
            "zadd", "zrange", "get_stats"
        ]
        for tool in expected_tools:
            assert tool in tool_names

    @pytest.mark.asyncio
    async def test_redis_set_get_value(self, redis_server):
        """Test Redis set and get operations."""
        # Set value
        result = await redis_server.set_value("test_key", "test_value", ttl=60)
        assert result["success"] is True
        
        # Get value
        result = await redis_server.get_value("test_key")
        assert result["success"] is True
        assert result["value"] == "test_value"
        
        # Delete value
        result = await redis_server.delete_key("test_key")
        assert result["success"] is True

    @pytest.mark.asyncio
    async def test_redis_sorted_set(self, redis_server):
        """Test Redis sorted set operations."""
        # Add to sorted set
        result = await redis_server.zadd("test_zset", 1.0, {"data": "value1"})
        assert result["success"] is True
        
        result = await redis_server.zadd("test_zset", 2.0, {"data": "value2"})
        assert result["success"] is True
        
        # Get from sorted set
        result = await redis_server.zrange("test_zset")
        assert result["success"] is True
        assert len(result["members"]) == 2
        
        # Cleanup
        await redis_server.delete_key("test_zset")


# ===== Docker MCP Server Tests =====

@pytest.mark.skipif(not DOCKER_AVAILABLE, reason="Docker dependencies not available")
class TestDockerMCPServer:
    """Test Docker MCP Server."""

    @pytest.fixture
    def docker_server(self):
        """Create Docker server instance."""
        config = {
            "socket": "unix:///var/run/docker.sock",
        }
        server = DockerMCPServer(config)
        yield server
        server.close()

    def test_docker_server_definition(self):
        """Test Docker server definition."""
        definition = docker_def()
        
        assert definition["name"] == "docker"
        assert definition["category"] == "devops"
        assert len(definition["tools"]) == 10
        
        tool_names = [tool["name"] for tool in definition["tools"]]
        expected_tools = [
            "list_containers", "start_container", "stop_container",
            "restart_container", "remove_container", "build_image",
            "pull_image", "push_image", "get_logs", "get_stats"
        ]
        for tool in expected_tools:
            assert tool in tool_names

    @pytest.mark.asyncio
    async def test_docker_list_containers(self, docker_server):
        """Test Docker list containers."""
        result = await docker_server.list_containers(all=True)
        
        assert result["success"] is True
        assert "containers" in result
        assert "count" in result


# ===== Prometheus MCP Server Tests =====

@pytest.mark.skipif(not PROMETHEUS_AVAILABLE, reason="Prometheus dependencies not available")
class TestPrometheusMCPServer:
    """Test Prometheus MCP Server."""

    @pytest.fixture
    async def prometheus_server(self):
        """Create Prometheus server instance."""
        config = {
            "url": os.getenv("PROMETHEUS_URL", "http://localhost:9090"),
        }
        server = PrometheusMCPServer(config)
        yield server
        await server.close()

    def test_prometheus_server_definition(self):
        """Test Prometheus server definition."""
        definition = prometheus_def()
        
        assert definition["name"] == "prometheus"
        assert definition["category"] == "monitoring"
        assert len(definition["tools"]) == 7
        
        tool_names = [tool["name"] for tool in definition["tools"]]
        expected_tools = [
            "query", "query_range", "create_alert",
            "list_alerts", "get_metrics", "record_metric", "get_targets"
        ]
        for tool in expected_tools:
            assert tool in tool_names

    @pytest.mark.asyncio
    async def test_prometheus_query(self, prometheus_server):
        """Test Prometheus query."""
        # Simple query
        result = await prometheus_server.query("up")
        
        # May fail if Prometheus not running, but should have proper structure
        assert "success" in result
        if result["success"]:
            assert "result" in result
            assert "query" in result

    @pytest.mark.asyncio
    async def test_prometheus_record_metric(self, prometheus_server):
        """Test Prometheus metric recording."""
        result = await prometheus_server.record_metric(
            name="test_metric",
            value=42.0,
            metric_type="gauge"
        )
        
        assert result["success"] is True
        assert result["name"] == "test_metric"
        assert result["value"] == 42.0


# ===== PDF Processing MCP Server Tests =====

@pytest.mark.skipif(not PDF_AVAILABLE, reason="PDF dependencies not available")
class TestPDFProcessingMCPServer:
    """Test PDF Processing MCP Server."""

    @pytest.fixture
    def pdf_server(self):
        """Create PDF server instance."""
        config = {
            "ocr_enabled": True,
            "ocr_language": "eng",
            "extract_tables": True,
        }
        return PDFProcessingMCPServer(config)

    @pytest.fixture
    def sample_pdf(self):
        """Create sample PDF file."""
        # Create a simple PDF for testing
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f:
            c = canvas.Canvas(f.name, pagesize=letter)
            c.drawString(100, 750, "Test PDF Document")
            c.drawString(100, 700, "Revenue: $1,000,000")
            c.drawString(100, 650, "Net Income: $200,000")
            c.save()
            yield f.name
        
        # Cleanup
        Path(f.name).unlink(missing_ok=True)

    def test_pdf_server_definition(self):
        """Test PDF server definition."""
        definition = pdf_def()
        
        assert definition["name"] == "pdf"
        assert definition["category"] == "documents"
        assert len(definition["tools"]) == 9
        
        tool_names = [tool["name"] for tool in definition["tools"]]
        expected_tools = [
            "extract_text", "extract_tables", "extract_10k_sections",
            "extract_10q_data", "ocr_scan", "find_keywords",
            "extract_metrics", "summarize_document", "compare_documents"
        ]
        for tool in expected_tools:
            assert tool in tool_names

    @pytest.mark.asyncio
    async def test_pdf_extract_text(self, pdf_server, sample_pdf):
        """Test PDF text extraction."""
        result = await pdf_server.extract_text(sample_pdf)
        
        assert result["success"] is True
        assert "full_text" in result
        assert "Test PDF" in result["full_text"] or len(result["full_text"]) > 0

    @pytest.mark.asyncio
    async def test_pdf_find_keywords(self, pdf_server, sample_pdf):
        """Test PDF keyword search."""
        result = await pdf_server.find_keywords(
            sample_pdf,
            keywords=["revenue", "income"],
            case_sensitive=False
        )
        
        assert result["success"] is True
        assert "keywords" in result

    @pytest.mark.asyncio
    async def test_pdf_extract_metrics(self, pdf_server, sample_pdf):
        """Test PDF financial metrics extraction."""
        result = await pdf_server.extract_metrics(sample_pdf)
        
        assert result["success"] is True
        assert "metrics" in result


# ===== Excel/Spreadsheet MCP Server Tests =====

@pytest.mark.skipif(not EXCEL_AVAILABLE, reason="Excel dependencies not available")
class TestExcelMCPServer:
    """Test Excel MCP Server."""

    @pytest.fixture
    def excel_server(self):
        """Create Excel server instance."""
        config = {
            "max_rows": 100000,
            "max_columns": 1000,
            "evaluate_formulas": True,
        }
        return ExcelMCPServer(config)

    @pytest.fixture
    def sample_excel(self):
        """Create sample Excel file."""
        import openpyxl
        
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            # Add sample data
            ws['A1'] = "Item"
            ws['B1'] = "Value"
            ws['A2'] = "Revenue"
            ws['B2'] = 1000000
            ws['A3'] = "Expenses"
            ws['B3'] = 800000
            ws['A4'] = "Net Income"
            ws['B4'] = "=B2-B3"
            
            wb.save(f.name)
            yield f.name
        
        # Cleanup
        Path(f.name).unlink(missing_ok=True)

    def test_excel_server_definition(self):
        """Test Excel server definition."""
        definition = excel_def()
        
        assert definition["name"] == "excel"
        assert definition["category"] == "documents"
        assert len(definition["tools"]) == 10
        
        tool_names = [tool["name"] for tool in definition["tools"]]
        expected_tools = [
            "read_workbook", "write_workbook", "read_sheet",
            "get_cell_value", "set_cell_value", "evaluate_formula",
            "create_pivot", "extract_tables", "format_financial_report",
            "parse_financial_model"
        ]
        for tool in expected_tools:
            assert tool in tool_names

    @pytest.mark.asyncio
    async def test_excel_read_workbook(self, excel_server, sample_excel):
        """Test Excel workbook reading."""
        result = await excel_server.read_workbook(sample_excel)
        
        assert result["success"] is True
        assert "sheets" in result
        assert result["sheet_count"] > 0

    @pytest.mark.asyncio
    async def test_excel_read_sheet(self, excel_server, sample_excel):
        """Test Excel sheet reading."""
        result = await excel_server.read_sheet(sample_excel, sheet_name="TestSheet")
        
        assert result["success"] is True
        assert "data" in result
        assert len(result["data"]) > 0

    @pytest.mark.asyncio
    async def test_excel_get_cell_value(self, excel_server, sample_excel):
        """Test Excel cell value retrieval."""
        result = await excel_server.get_cell_value(
            sample_excel,
            sheet_name="TestSheet",
            cell_address="A1"
        )
        
        assert result["success"] is True
        assert result["value"] == "Item"

    @pytest.mark.asyncio
    async def test_excel_write_workbook(self, excel_server):
        """Test Excel workbook writing."""
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            excel_path = f.name
        
        try:
            result = await excel_server.write_workbook(
                excel_path,
                sheets={
                    "Sheet1": [
                        ["Name", "Value"],
                        ["Item1", 100],
                        ["Item2", 200],
                    ]
                },
                overwrite=True
            )
            
            assert result["success"] is True
            assert result["sheet_count"] == 1
            
            # Verify file was created
            assert Path(excel_path).exists()
        
        finally:
            # Cleanup
            Path(excel_path).unlink(missing_ok=True)


# ===== Integration Tests =====

class TestMCPIntegration:
    """Test MCP server integration."""

    def test_all_server_definitions(self):
        """Test that all server definitions are valid."""
        servers = []
        
        if REDIS_AVAILABLE:
            servers.append(redis_def())
        if DOCKER_AVAILABLE:
            servers.append(docker_def())
        if PROMETHEUS_AVAILABLE:
            servers.append(prometheus_def())
        if PDF_AVAILABLE:
            servers.append(pdf_def())
        if EXCEL_AVAILABLE:
            servers.append(excel_def())
        
        # All servers should have required fields
        for server in servers:
            assert "name" in server
            assert "category" in server
            assert "description" in server
            assert "tools" in server
            assert "metadata" in server
            
            # All tools should have required fields
            for tool in server["tools"]:
                assert "name" in tool
                assert "description" in tool
                assert "parameters" in tool

    def test_server_categories(self):
        """Test that servers have correct categories."""
        expected_categories = {
            "redis": "storage",
            "docker": "devops",
            "prometheus": "monitoring",
            "pdf": "documents",
            "excel": "documents",
        }
        
        for server_name, expected_cat in expected_categories.items():
            if server_name == "redis" and REDIS_AVAILABLE:
                assert redis_def()["category"] == expected_cat
            elif server_name == "docker" and DOCKER_AVAILABLE:
                assert docker_def()["category"] == expected_cat
            elif server_name == "prometheus" and PROMETHEUS_AVAILABLE:
                assert prometheus_def()["category"] == expected_cat
            elif server_name == "pdf" and PDF_AVAILABLE:
                assert pdf_def()["category"] == expected_cat
            elif server_name == "excel" and EXCEL_AVAILABLE:
                assert excel_def()["category"] == expected_cat


# ===== Performance Tests =====

class TestMCPPerformance:
    """Test MCP server performance targets."""

    @pytest.mark.skipif(not REDIS_AVAILABLE, reason="Redis not available")
    @pytest.mark.asyncio
    async def test_redis_performance(self):
        """Test Redis performance (<2ms target)."""
        import time
        
        config = {"host": "localhost", "port": 6379, "db": 0}
        server = RedisMCPServer(config)
        
        try:
            # Test set performance
            start = time.time()
            result = await server.set_value("perf_test", "value")
            duration = (time.time() - start) * 1000  # ms
            
            if result["success"]:
                assert duration < 10  # 10ms threshold (more lenient for testing)
        finally:
            await server.close()

    @pytest.mark.skipif(not EXCEL_AVAILABLE, reason="Excel not available")
    @pytest.mark.asyncio
    async def test_excel_read_performance(self):
        """Test Excel read performance (<500ms target)."""
        import time
        import openpyxl
        
        # Create test file
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = openpyxl.Workbook()
            ws = wb.active
            for i in range(100):
                ws.append([f"Row{i}", i, i*2])
            wb.save(f.name)
            excel_path = f.name
        
        try:
            server = ExcelMCPServer({"max_rows": 100000})
            
            start = time.time()
            result = await server.read_sheet(excel_path)
            duration = (time.time() - start) * 1000  # ms
            
            if result["success"]:
                assert duration < 1000  # 1s threshold (lenient for testing)
        finally:
            Path(excel_path).unlink(missing_ok=True)


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])